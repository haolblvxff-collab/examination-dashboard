#!/usr/bin/env python3
"""运城培优成绩追踪看板 - FastAPI 后端
数据解析 + matplotlib 图表生成 + 模板下载
"""

import io, os, json, base64, sys
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
from fastapi import FastAPI, UploadFile, File, Query
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = FastAPI(title="运城培优成绩追踪")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── 中文字体（跨平台自适应，避免冻结环境扫描卡死）──
import platform as _platform
_sys = _platform.system()
if _sys == 'Darwin':
    plt.rcParams['font.sans-serif'] = ['PingFang SC', 'STHeiti', 'Heiti TC', 'DejaVu Sans']
elif _sys == 'Windows':
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
else:
    plt.rcParams['font.sans-serif'] = ['WenQuanYi Micro Hei', 'Noto Sans CJK SC', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

BASE = Path(__file__).parent
STATIC = BASE / "static"

SUBJECTS = ['语文', '数学', '英语', '物理', '化学', '生物']
EXAM_ORDER = ['一模', '二模', '三模', '四模', '五模']
COLORS = {'语文': '#ef4444', '数学': '#f59e0b', '英语': '#10b981',
          '物理': '#3b82f6', '化学': '#8b5cf6', '生物': '#06b6d4'}
PALETTE = ['#ef4444', '#f59e0b', '#10b981', '#3b82f6', '#8b5cf6', '#06b6d4',
           '#2563eb', '#7c3aed', '#ec4899', '#14b8a6']

DATA = {"students": {}, "exams": [], "schools": [], "raw": None}

def dark_style():
    plt.style.use('dark_background')
    sns.set_palette(PALETTE)

# ── 解析上传的 Excel ──
def parse_excel(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("文件至少需要表头 + 1 行数据")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = rows[1:]
    df = pd.DataFrame(data_rows, columns=headers)
    df = df.dropna(how='all')

    name_col = next((c for c in df.columns if any(k in str(c) for k in ['姓名', '名字', '学生'])), None)
    sch_col = next((c for c in df.columns if '学校' in str(c)), None)
    exam_col = next((c for c in df.columns if any(k in str(c) for k in ['考试', '场次'])), None)

    if not name_col:
        raise ValueError("找不到「姓名」列，请使用模板格式")

    exams = []
    students = {}

    if exam_col:
        exam_vals = df[exam_col].dropna().unique()
        exams = sorted(exam_vals, key=lambda x: EXAM_ORDER.index(x) if x in EXAM_ORDER else 99)
        subjects_found = [c for c in df.columns if c in SUBJECTS]
        for _, row in df.iterrows():
            name = str(row[name_col]).strip()
            if not name or name == 'nan':
                continue
            sch = str(row[sch_col]).strip() if sch_col and pd.notna(row[sch_col]) else '—'
            exam = str(row[exam_col]).strip()
            if name not in students:
                students[name] = {'school': sch, 'scores': {}, 'totals': {}}
            total = 0
            for sub in subjects_found:
                v = row[sub]
                if pd.notna(v) and isinstance(v, (int, float)):
                    if sub not in students[name]['scores']:
                        students[name]['scores'][sub] = {}
                    students[name]['scores'][sub][exam] = float(v)
                    total += float(v)
            if total > 0:
                students[name]['totals'][exam] = round(total, 1)
    else:
        exam_map = {}
        for c in df.columns:
            parts = str(c).split('_', 1)
            if len(parts) == 2 and parts[1] in SUBJECTS:
                exam_map.setdefault(parts[0], {})[parts[1]] = str(c)
        exams = sorted(exam_map.keys(), key=lambda x: EXAM_ORDER.index(x) if x in EXAM_ORDER else 99)

        for _, row in df.iterrows():
            name = str(row[name_col]).strip()
            if not name or name == 'nan':
                continue
            sch = str(row[sch_col]).strip() if sch_col and pd.notna(row[sch_col]) else '—'
            if name not in students:
                students[name] = {'school': sch, 'scores': {}, 'totals': {}}
            for exam in exams:
                total = 0
                for sub in SUBJECTS:
                    col = exam_map.get(exam, {}).get(sub)
                    if col and col in df.columns:
                        v = row[col]
                        if pd.notna(v) and isinstance(v, (int, float)):
                            if sub not in students[name]['scores']:
                                students[name]['scores'][sub] = {}
                            students[name]['scores'][sub][exam] = float(v)
                            total += float(v)
                if total > 0:
                    students[name]['totals'][exam] = round(total, 1)

    schools = sorted(set(s['school'] for s in students.values()))
    return {"students": students, "exams": exams, "schools": schools}


# ═══════════════════════════════════════════
# API Routes
# ═══════════════════════════════════════════

@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    try:
        content = await file.read()
        result = parse_excel(content)
        DATA.update(result)
        DATA['raw'] = content
        DATA['filename'] = file.filename
        return {"ok": True, "students": len(result['students']), "exams": len(result['exams']),
                "schools": len(result['schools']), "filename": file.filename}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.get("/api/data")
def get_data():
    if not DATA.get('students'):
        return {"ok": False, "error": "请先上传成绩文件"}
    return {"ok": True, "students": DATA['students'], "exams": DATA['exams'],
            "schools": DATA['schools'], "filename": DATA.get('filename', ''), "subjects": SUBJECTS}


@app.get("/api/stats")
def get_stats():
    if not DATA.get('students'):
        return JSONResponse({"error": "请先上传成绩文件"}, status_code=400)
    s = DATA['students']
    exams = DATA['exams']
    last_exam = exams[-1] if exams else None

    names = [n for n in s if last_exam and s[n]['totals'].get(last_exam)]
    n_students = len(s)
    n_active = len(names)

    max_total, max_name, avg_total = 0, '', 0
    for n in names:
        t = s[n]['totals'].get(last_exam, 0)
        avg_total += t
        if t > max_total:
            max_total, max_name = t, n
    avg_total = round(avg_total / n_active, 1) if n_active else 0

    best_improve, best_improve_info = 0, ''
    if len(exams) >= 2:
        for n in names:
            for i in range(len(exams) - 1, 0, -1):
                cur = s[n]['totals'].get(exams[i])
                prev = s[n]['totals'].get(exams[i - 1])
                if cur and prev and cur - prev > best_improve:
                    best_improve = round(cur - prev, 1)
                    best_improve_info = f"{n} ({exams[i-1]}→{exams[i]})"
                break

    subject_avg = {}
    for sub in SUBJECTS:
        scores = [s[n]['scores'].get(sub, {}).get(last_exam) for n in names
                  if s[n]['scores'].get(sub, {}).get(last_exam) is not None]
        subject_avg[sub] = round(np.mean(scores), 1) if scores else 0

    return {"n_students": n_students, "n_exams": len(exams), "n_schools": len(DATA['schools']),
            "max_total": max_total, "max_name": max_name, "avg_total": avg_total,
            "best_improve": best_improve, "best_improve_info": best_improve_info,
            "subject_avg": subject_avg, "exams": exams}


@app.get("/api/ranking")
def ranking(exam: str = Query(...), school: str = Query("")):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    s = DATA['students']
    items = []
    for name, info in s.items():
        if school and info['school'] != school:
            continue
        t = info['totals'].get(exam)
        if t is None:
            continue
        row = {"name": name, "school": info['school'], "total": t}
        for sub in SUBJECTS:
            row[sub] = info['scores'].get(sub, {}).get(exam)
        items.append(row)
    items.sort(key=lambda x: x['total'], reverse=True)
    return {"exam": exam, "ranking": items}


# ═══════════════════════════════════════════
# 图表生成 (matplotlib)
# ═══════════════════════════════════════════

def _save_chart(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=140, bbox_inches='tight', facecolor='#0f172a')
    buf.seek(0)
    plt.close(fig)
    return StreamingResponse(buf, media_type="image/png")


@app.get("/api/chart/trend")
def chart_trend(view: str = "exam", subject: str = "", student: str = "",
                school: str = "", school_student: str = ""):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    exams = DATA['exams']
    if not exams:
        return JSONResponse({"error": "无考试数据"}, status_code=400)

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')

    if view == 'exam':
        ns = [n for n in s if not school or s[n]['school'] == school]
        avgs = [np.mean([s[n]['totals'][ex] for n in ns if s[n]['totals'].get(ex)]) if 
                [s[n]['totals'][ex] for n in ns if s[n]['totals'].get(ex)] else None for ex in exams]
        ax.plot(range(len(exams)), avgs, color='#38bdf8', linewidth=3, marker='o', markersize=8)
        ax.set_title('全校总分均分走势', color='white', fontsize=14, pad=15)
        ax.set_ylabel('总分', color='#94a3b8')

    elif view == 'subject':
        sub_list = [subject] if subject else SUBJECTS
        ns = [n for n in s if not school or s[n]['school'] == school]
        for sub in sub_list:
            avgs = []
            for ex in exams:
                vals = [s[n]['scores'].get(sub, {}).get(ex) for n in ns
                        if s[n]['scores'].get(sub, {}).get(ex) is not None]
                avgs.append(np.mean(vals) if vals else None)
            ax.plot(range(len(exams)), avgs, color=COLORS.get(sub, '#888'),
                    linewidth=2.5, marker='o', markersize=6, label=sub)
        ax.set_title('各科均分走势', color='white', fontsize=14, pad=15)
        ax.legend(loc='upper left', fontsize=10)
        ax.set_ylabel('分数', color='#94a3b8')

    elif view == 'student':
        if not student:
            return JSONResponse({"error": "请选择学生"}, status_code=400)
        info = s.get(student)
        if not info:
            return JSONResponse({"error": "学生不存在"}, status_code=400)
        for sub in SUBJECTS:
            vals = [info['scores'].get(sub, {}).get(ex) for ex in exams]
            ax.plot(range(len(exams)), vals, color=COLORS.get(sub, '#888'),
                    linewidth=2, marker='s', markersize=5, label=sub)
        totals = [info['totals'].get(ex) for ex in exams]
        ax.plot(range(len(exams)), totals, color='white', linewidth=3,
                marker='D', markersize=7, linestyle='--', label='总分')
        ax.set_title(f'{student} 各科走势', color='white', fontsize=14, pad=15)
        ax.legend(loc='upper left', fontsize=10)
        ax.set_ylabel('分数', color='#94a3b8')

    elif view == 'school':
        if school_student:
            info = s.get(school_student)
            if not info:
                return JSONResponse({"error": "学生不存在"}, status_code=400)
            for sub in SUBJECTS:
                vals = [info['scores'].get(sub, {}).get(ex) for ex in exams]
                ax.plot(range(len(exams)), vals, color=COLORS.get(sub, '#888'),
                        linewidth=2, marker='s', markersize=5, label=sub)
            ax.set_title(f'{school_student} 各科走势', color='white', fontsize=14, pad=15)
            ax.legend(loc='upper left', fontsize=10)
        elif school:
            for sub in SUBJECTS:
                ns = [n for n in s if s[n]['school'] == school]
                avgs = []
                for ex in exams:
                    vals = [s[n]['scores'].get(sub, {}).get(ex) for n in ns
                            if s[n]['scores'].get(sub, {}).get(ex) is not None]
                    avgs.append(np.mean(vals) if vals else None)
                ax.plot(range(len(exams)), avgs, color=COLORS.get(sub, '#888'),
                        linewidth=2.5, marker='o', markersize=6, label=sub)
            ax.set_title(f'{school} 各科均分走势', color='white', fontsize=14, pad=15)
            ax.legend(loc='upper left', fontsize=10)
        else:
            for i, sch in enumerate(DATA['schools']):
                ns = [n for n in s if s[n]['school'] == sch]
                avgs = []
                for ex in exams:
                    vals = [s[n]['totals'][ex] for n in ns if s[n]['totals'].get(ex)]
                    avgs.append(np.mean(vals) if vals else None)
                ax.plot(range(len(exams)), avgs, color=PALETTE[i % len(PALETTE)],
                        linewidth=2.5, marker='o', markersize=6, label=f'{sch}({len(ns)}人)')
            ax.set_title('各校总分均分对比', color='white', fontsize=14, pad=15)
            ax.legend(loc='upper left', fontsize=10)
        ax.set_ylabel('分数', color='#94a3b8')

    ax.set_xticks(range(len(exams)))
    ax.set_xticklabels(exams, color='#94a3b8', fontsize=11)
    ax.tick_params(colors='#94a3b8')
    ax.grid(axis='y', color='#334155', alpha=0.6)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_color('#334155')
    ax.spines['left'].set_color('#334155')
    return _save_chart(fig)


@app.get("/api/chart/distribution")
def chart_distribution(exam: str = Query(...), subject: str = Query("")):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    sub_list = [subject] if subject else SUBJECTS
    n = len(sub_list)
    cols = min(n, 3)
    rows = (n + cols - 1) // cols

    fig, axes = plt.subplots(rows, cols, figsize=(cols * 4.5, rows * 3.5))
    fig.patch.set_facecolor('#0f172a')
    axes = [axes] if n == 1 else axes.flatten()

    for i, sub in enumerate(sub_list):
        ax = axes[i]
        ax.set_facecolor('#1e293b')
        scores = []
        for name, info in s.items():
            sc = info['scores'].get(sub, {}).get(exam)
            if sc is not None:
                scores.append(sc)
        if scores:
            ax.hist(scores, bins=15, color=COLORS.get(sub, '#38bdf8'), alpha=0.8, edgecolor='white', linewidth=0.5)
            mean_v = np.mean(scores)
            ax.axvline(mean_v, color='white', linestyle='--', linewidth=1.5, label=f'均分 {mean_v:.1f}')
            ax.legend(fontsize=9)
        ax.set_title(f'{sub} 分数分布 ({exam})', color='white', fontsize=11)
        ax.tick_params(colors='#94a3b8')
        for spine in ['top', 'right']:
            ax.spines[spine].set_visible(False)
        ax.spines['bottom'].set_color('#334155')
        ax.spines['left'].set_color('#334155')
        ax.grid(axis='y', color='#334155', alpha=0.5)

    for i in range(n, len(axes)):
        axes[i].set_visible(False)
    return _save_chart(fig)


@app.get("/api/chart/radar")
def chart_radar(student: str = Query(...), exam: str = Query(...)):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    info = DATA['students'].get(student)
    if not info:
        return JSONResponse({"error": "学生不存在"}, status_code=400)

    values = [info['scores'].get(sub, {}).get(exam) or 0 for sub in SUBJECTS]
    angles = np.linspace(0, 2 * np.pi, len(SUBJECTS), endpoint=False).tolist()
    values += values[:1]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')
    ax.fill(angles, values, color='#38bdf8', alpha=0.3)
    ax.plot(angles, values, color='#38bdf8', linewidth=2.5, marker='o', markersize=8)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(SUBJECTS, color='white', fontsize=12)
    ax.set_ylim(0, max(max(values) * 1.2, 100))
    ax.set_yticks([])
    ax.set_title(f'{student} · {exam} 学科雷达', color='white', fontsize=14, pad=20)
    ax.spines['polar'].set_color('#334155')
    ax.grid(color='#475569', alpha=0.6)
    return _save_chart(fig)


@app.get("/api/chart/comparison")
def chart_comparison(exam: str = Query(...)):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    schools = DATA['schools']
    if not schools:
        return JSONResponse({"error": "无学校数据"}, status_code=400)

    fig, ax = plt.subplots(figsize=(10, 5.5))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')

    x = np.arange(len(SUBJECTS))
    width = 0.8 / len(schools)
    for i, sch in enumerate(schools):
        ns = [n for n in s if s[n]['school'] == sch]
        avgs = []
        for sub in SUBJECTS:
            vals = [s[n]['scores'].get(sub, {}).get(exam) for n in ns
                    if s[n]['scores'].get(sub, {}).get(exam) is not None]
            avgs.append(np.mean(vals) if vals else 0)
        offset = (i - len(schools) / 2 + 0.5) * width
        ax.bar(x + offset, avgs, width, label=f'{sch}({len(ns)}人)',
               color=PALETTE[i % len(PALETTE)], alpha=0.85, edgecolor='white', linewidth=0.3)

    ax.set_xticks(x)
    ax.set_xticklabels(SUBJECTS, color='white', fontsize=12)
    ax.set_title(f'各校各科均分对比 ({exam})', color='white', fontsize=14, pad=15)
    ax.set_ylabel('均分', color='#94a3b8')
    ax.legend(fontsize=10, loc='upper right')
    ax.tick_params(colors='#94a3b8')
    ax.grid(axis='y', color='#334155', alpha=0.5)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    ax.spines['bottom'].set_color('#334155')
    ax.spines['left'].set_color('#334155')
    return _save_chart(fig)


@app.get("/api/chart/heatmap")
def chart_heatmap(exam: str = Query(...)):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    names = sorted([n for n in s if s[n]['totals'].get(exam)],
                   key=lambda n: s[n]['totals'].get(exam, 0), reverse=True)
    if not names:
        return JSONResponse({"error": "无有效数据"}, status_code=400)

    matrix = np.array([[s[name]['scores'].get(sub, {}).get(exam) 
                        if s[name]['scores'].get(sub, {}).get(exam) is not None else np.nan
                        for sub in SUBJECTS] for name in names])
    h = max(4, len(names) * 0.35)
    fig, ax = plt.subplots(figsize=(10, h))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')

    im = ax.imshow(matrix, cmap='YlOrRd', aspect='auto', vmin=np.nanmin(matrix), vmax=np.nanmax(matrix))
    ax.set_xticks(range(len(SUBJECTS)))
    ax.set_xticklabels(SUBJECTS, color='white', fontsize=11)
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, color='#94a3b8', fontsize=9)
    ax.set_title(f'学生 × 科目 成绩热力图 ({exam})', color='white', fontsize=14, pad=15)
    plt.setp(ax.get_xticklabels(), rotation=0, ha="center")

    cbar = fig.colorbar(im, ax=ax, shrink=0.8)
    cbar.ax.yaxis.set_tick_params(color='white')
    plt.setp(plt.getp(cbar.ax.axes, 'yticklabels'), color='white')

    for i in range(len(names)):
        for j in range(len(SUBJECTS)):
            v = matrix[i, j]
            if not np.isnan(v):
                ax.text(j, i, f'{v:.0f}', ha='center', va='center',
                        color='white' if v < np.nanmean(matrix) + 10 else '#1e293b',
                        fontsize=8, fontweight='bold')
    return _save_chart(fig)


@app.get("/api/chart/boxplot")
def chart_boxplot(exam: str = Query(...)):
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    data, labels = [], []
    for sub in SUBJECTS:
        vals = []
        for name, info in s.items():
            sc = info['scores'].get(sub, {}).get(exam)
            if sc is not None:
                vals.append(sc)
        if vals:
            data.append(vals)
            labels.append(sub)

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')
    ax.boxplot(data, labels=labels, patch_artist=True,
               medianprops=dict(color='white', linewidth=2),
               whiskerprops=dict(color='#94a3b8'),
               capprops=dict(color='#94a3b8'),
               boxprops=dict(facecolor='#3b82f6', alpha=0.6, edgecolor='#94a3b8'),
               flierprops=dict(markerfacecolor='#ef4444', marker='o', markersize=6))
    ax.set_title(f'各科分数箱线图 ({exam})', color='white', fontsize=14, pad=15)
    ax.set_ylabel('分数', color='#94a3b8')
    ax.tick_params(colors='#94a3b8')
    ax.grid(axis='y', color='#334155', alpha=0.5)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    ax.spines['bottom'].set_color('#334155')
    ax.spines['left'].set_color('#334155')
    return _save_chart(fig)


@app.get("/api/chart/improvement")
def chart_improvement():
    if not DATA.get('students'):
        return JSONResponse({"error": "无数据"}, status_code=400)
    dark_style()
    s = DATA['students']
    exams = DATA['exams']
    if len(exams) < 2:
        return JSONResponse({"error": "需要至少2场考试"}, status_code=400)

    first_exam, last_exam = exams[0], exams[-1]
    names, changes = [], []
    for name, info in s.items():
        first = info['totals'].get(first_exam)
        last = info['totals'].get(last_exam)
        if first and last:
            names.append(name)
            changes.append(round(last - first, 1))

    if not names:
        return JSONResponse({"error": "无有效数据"}, status_code=400)

    sorted_idx = np.argsort(changes)
    names = [names[i] for i in sorted_idx]
    changes = [changes[i] for i in sorted_idx]
    colors = ['#10b981' if c >= 0 else '#ef4444' for c in changes]

    fig, ax = plt.subplots(figsize=(8, max(4, len(names) * 0.4)))
    fig.patch.set_facecolor('#0f172a')
    ax.set_facecolor('#1e293b')
    bars = ax.barh(range(len(names)), changes, color=colors, alpha=0.85, height=0.7)
    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, color='#94a3b8', fontsize=10)
    ax.axvline(0, color='white', linewidth=0.8)
    ax.set_title(f'总分进步幅度 ({first_exam} → {last_exam})', color='white', fontsize=14, pad=15)
    ax.set_xlabel('分数变化', color='#94a3b8')
    ax.tick_params(colors='#94a3b8')
    ax.grid(axis='x', color='#334155', alpha=0.5)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    ax.spines['bottom'].set_color('#334155')
    ax.spines['left'].set_color('#334155')

    for bar, val in zip(bars, changes):
        ax.text(val + (0.5 if val >= 0 else -0.5), bar.get_y() + bar.get_height() / 2,
                f'{val:+.1f}', va='center', ha='left' if val >= 0 else 'right',
                color='white', fontsize=9)
    return _save_chart(fig)


# ═══════════════════════════════════════════
# 模板下载
# ═══════════════════════════════════════════

@app.get("/api/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩录入"

    header_font = Font(name='PingFang SC', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'),
                         top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
    data_font = Font(name='PingFang SC', size=11)
    data_align = Alignment(horizontal='center', vertical='center')

    headers = ['姓名', '学校', '考试'] + SUBJECTS + ['总分(自动计算)', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    demo_data = [
        ['张三', '运城培优', '一模', 108, 120, 115, 85, 78, 82, None, '示例-可删除'],
        ['张三', '运城培优', '二模', 112, 125, 118, 88, 82, 85, None, ''],
        ['李四', '运城二中', '一模', 98, 110, 105, 78, 72, 76, None, ''],
    ]
    for r, row_data in enumerate(demo_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    col_widths = [12, 14, 8, 8, 8, 8, 8, 8, 8, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws2 = wb.create_sheet("使用说明")
    notes = [
        ["📋 成绩录入模板使用说明"], [""],
        ["1. 请在「成绩录入」sheet 中填写数据"],
        ["2. 必填列：姓名、学校、考试 —— 三者共同标识一条记录"],
        ["3. 同一学生多次考试需重复填写姓名+学校，考试列写不同场次"],
        ["4. 考试名称自由填写，系统自动识别排序"],
        ["5. 分数填写数字即可，未参加科目留空"],
        ["6. 总分列无需手动填写，系统自动计算"],
        ["7. 示例数据可直接删除，填写真实数据后上传"],
        ["8. 上传格式：.xlsx / .xls"],
        [""], ["✅ 推荐：长格式（每行=学生+考试+各科分数）"],
    ]
    for r, row in enumerate(notes, 1):
        cell = ws2.cell(row=r, column=1, value=row[0])
        cell.font = Font(name='PingFang SC', size=14 if r == 1 else 11, bold=(r == 1),
                         color='2563EB' if r == 1 else '374151')
    ws2.column_dimensions['A'].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    encoded_filename = quote('成绩录入模板.xlsx')
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"})


# ═══════════════════════════════════════════
# 首页
# ═══════════════════════════════════════════

def _get_resource_path(relative_path: str) -> Path:
    if getattr(sys, 'frozen', False):
        base = Path(sys._MEIPASS)
    else:
        base = BASE
    return base / relative_path

_HTML_TEMPLATE = None

def _load_html() -> str:
    global _HTML_TEMPLATE
    if _HTML_TEMPLATE is None:
        html_path = _get_resource_path("templates/index.html")
        _HTML_TEMPLATE = html_path.read_text(encoding='utf-8')
    return _HTML_TEMPLATE

@app.get("/", response_class=HTMLResponse)
def index():
    return _load_html()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8899)
